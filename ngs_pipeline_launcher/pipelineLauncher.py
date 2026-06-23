import pandas as pd, os, search_tools as st, shutil, io, time, subprocess, argparse, tempfile, pathlib, glob, re, glob, re
from configparser import ConfigParser
import openpyxl as xl
from runStatus import *
pd.options.mode.chained_assignment = None  # default='warn'
os.chdir(os.path.dirname(__file__))

defaultSampleSheet = "./"
SLURM = "/nfs/APL_Genomics/apps/production/ngs-pipeline-launcher/templates/SLURM_template.batch"
barcodeCol = "Barcode" # The column in the Pipeline Worksheet [Samples] section that contains the barcode
samplePosCol = "Sample_Pos" # Generated column in the Pipeline Worksheet [Samples] section that contains the sample name (e.g., "27_S" for illumina)
symlinkFQ = False # Should fastq's be symlinked?
pathfilter = ["**/*fastq.gz","**/*fast5","**/report_*.json","**/*PipelineWorksheet*"]  # Which files should be transferred?

def getSampleSheetDataVars(path:str, section:str):
    """Generates a dictionary from the first two columns of a [HEADER] section
    :param path: The path to the reference file
    :param section: The name of the section
    :return: A dictionary containing variable keys
    """    
    cfg = ConfigParser(allow_no_value=True)
    cfg.optionxform = str
    cfg.read(path)
    dict = {k:v for k, *v in map(lambda x: str.split(x,sep=","), cfg[section])}
    dict = {k:v[0] for k, v in dict.items() if v} # Removes blank keys and keep only first column after var
    return (dict)

def getSampleSheetDataFrame(path:str, section:str):
    """Generates a DataFrame from a [HEADER] section
    :param path: The path to the reference file
    :param section: The name of the section
    :return: A DataFrame representing the sections
    """ 
    cfg = ConfigParser(allow_no_value=True)
    cfg.optionxform = str
    cfg.read(path)
    buf = io.StringIO()
    buf.writelines('\n'.join(row.rstrip(',') for row in cfg[section]))
    buf.seek(0)
    df = pd.read_csv(buf)
    return (df)

def subsetWorksheet(path:str, group:str, out_path:str):
    """Subsets a worksheet to only include a specific Sample_Group  in the [SAMPLES] section
    :param group: The group to look for in the column 'Sample_Group'
    :param path: The output path of the subsetted pipeline worksheet
    :return: A DataFrame representing the sections
    """ 
    wb = xl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
    for row in reversed(rows): 
        cell = row[2] # col idx 3 is Sample_Group
        if cell.value == "Sample_Group":
            break
        if cell.value != group:
            ws.delete_rows(cell.row, 1)
    wb.save(out_path)

def generateSLURM(SLURM:str, jobName: str, runName: str, outputDir: str, command: str, email: str = None):
    """Generates a SLURM command file based on a template
    :param SLURM: Path to the template SLURM file
    :param jobName: The name of the job
    :param outputDir: The directory for output/error files
    :param command: The path to the new SLURM file
    """
    file = open(SLURM, "rt")
    data = file.read()
    file.close()
    data = data.replace("[JOB_NAME]", jobName)
    data = data.replace("[OUTPUT_DIR]", os.path.join(outputDir,runName+"_out.txt"))
    data = data.replace("[ERROR_DIR]", os.path.join(outputDir,runName+"_error.txt"))
    data = data.replace("[EMAIL]",  "NONE" if email is None else email)
    data = data.replace("[MAIL_TYPE]", "NONE" if email is None else "ALL")      
    data = data.replace("[RUN_DIR]", outputDir)
    outFile = os.path.join(outputDir,runName+"_SLURM.batch")
    file = open(outFile, "wt+")
    file.write(data)
    file.write(f"\n\ncd {outputDir}")
    file.write("\n\n"+command)
    file.close()
    return(outFile)

def printLog (message: str) :
    """Prints a message formatted as "[Current time] | [Message]
    :param message: The message to print
    """
    print (f"{currentTime()} | {message}", flush=True)

def runLauncher(sampleSheetPath: str, email: str = None, force = False):
    """Generates a SLURM command file based on a template
    :param sampleSheetPath: The path to the directory containing the sample sheet. The file must contain the sub-string 'PipelineWorksheet'.
    :param email: If desired, SLURM will send an e-mail on job status. Default = None
    :param force: Whether to delete the target directories, if they exist. Should only used while developing/debugging. Default = False
    """
    printLog(f"Pipeline launcher initialized in '{sampleSheetPath}'...")

    # Read data from the sample sheet
    printLog(f"Checking for pipeline worksheet...")
    with tempfile.NamedTemporaryFile() as sampleSheet:

        # Check for either specific file or directory to search
        if os.path.isfile(sampleSheetPath): # If file
            if os.path.exists(sampleSheetPath):
                file = sampleSheetPath  
            else:
                raise Exception(f"Pipeline worksheet not found. Please check '{sampleSheetPath}'.")
        elif os.path.isdir(sampleSheetPath): # If directory, then search for worksheet
            file = st.findFiles2(os.path.join(sampleSheetPath,"**","*PipelineWorksheet*"))
            if not isinstance(file, list): file = [file]
            file = [ f for f in file if "~$" not in f ] # Exclude temp files (e.g., if open in Excel)
            if (len(file) == 0): # If no files found
                raise Exception(f"No pipeline worksheet found. Please check '{sampleSheetPath}'.")
            if (len(file) > 1): # If more than 1 file is found
                raise Exception(f"More than one pipeline worksheet identified. Found:\n{file}.")
            file = file[0]
        
        printLog(f"   Found '{file}'")
        
        # Convert to CSV
        if pathlib.Path(file).suffix.lower() == ".xlsx":
            df = pd.read_excel(file)
            sampleSheetPath = sampleSheet.name # Export df to the tempfile
            df.to_csv(sampleSheetPath, index=False)
        elif pathlib.Path(file).suffix.lower() == ".csv": 
            sampleSheetPath = file
        else:
            raise Exception(f"Pipeline worksheet must have the file type of '.xlsx' or '.csv'. Please check '{file}'.")

        # Get the variables for the run
        header = getSampleSheetDataVars(sampleSheetPath, "Header") 
        runName = header["Run_Name"].strip()
        runDir = header["Run_Dir"].strip()
            
        pipelines = getSampleSheetDataVars(sampleSheetPath, "Pipelines")
        directories = getSampleSheetDataVars(sampleSheetPath, "Directories")   
        directories = {group: os.path.join(dir,runName) for group, dir in directories.items()} # Adds path name to end of directory path
        allSamples = getSampleSheetDataFrame(sampleSheetPath, "Samples")

    # Check if run is finished sequencing
    if not os.path.isdir(runDir): # Check if run exists
        raise Exception(f"Run directory does not exist at '{header['Run_Dir']}'")

    printLog(f"Checking for sequencing completion file...")
    sleep_time = 60
    while not (completionFiles := isRunCompleted(runDir)):#isRunCompleted(basePath, header["Seq_Type"]):
        printLog(f"   Waiting...")
        time.sleep(sleep_time)
        sleep_time = min(3600, sleep_time*2)

    printLog(f"   Found '{completionFiles[0]}'")

    # Check for appropriate inputs
    groups = sorted(set(allSamples["Sample_Group"].dropna().values))
    for group in groups:

        if group == "ignore":
            continue
        
        if not os.path.exists(pipelines[group]): # Check if pipeline exists
            if (pipelines[group] != "ignore"):
                raise Exception(f"Pipeline for '{group}' does not exist at '{pipelines[group]}'")

        if os.path.exists(directories[group]): # Check if directories exist
            if (directories[group] != "ignore"):
                if (len(directories[group]) != 0):
                    if not force:
                        raise Exception(f"Directory for '{group}' at '{directories[group]}' already exists and is not empty. Please choose empty or non-existing directory.")
                    else:
                        printLog(f"Removing dir: {directories[group]}.")
                        shutil.rmtree(directories[group], ignore_errors=True)

    # time.sleep(15*60) # Extra wait to make sure everything is done

    # Add barcodes to the respective sequencing type
    allSamples[samplePosCol] = allSamples[barcodeCol]
    allBarcodes = range(1,1000) # 1000 is arbitrary. Only needs to be higher than the max barcode value.

    if ("_I_" in runName):
        label = lambda x: f"{x}_S"
    else:
        label = lambda x: f"barcode{x}" if int(x) >= 10 else f"barcode0{x}"
        
    allSamples[barcodeCol] = allSamples[barcodeCol].apply(label)
    allBarcodes = [label(barcode) for barcode in allBarcodes]

    # Split sequencing run into respective folders
    printLog(f"Locating for files to move...")

    for group in groups:
        if group.lower() == "ignore": continue

        # Check for directory
        outDir = directories[group]
        ignore = False
        try: ignore = outDir.lower() == "ignore"
        except KeyError: ignore = True
        if (ignore): printLog(f"   Ignoring file copy for {group}"); continue

        # Get barcodes to include
        printLog(f"   Moving {group} to '{outDir}'")
        includeSamples = allSamples.loc[allSamples['Sample_Group'] == group][barcodeCol].values.tolist()
        includeSamples = st.sortDigitSuffix(list(includeSamples))
        printLog(f"      Extracting barcodes: " + ", ".join(st.collapseNumbers(includeSamples)))

        # Get barcodes to exclude
        excludeSamples = list(set(allBarcodes) - set(includeSamples))
        excludeSamples = st.sortDigitSuffix(list(excludeSamples))
        # print("      Excluding barcodes: " + ", ".join(st.collapseNumbers(excludeSamples)))
        excludeSamples = [f"\/{sample}|\/.*_{sample}|\/.*-{sample}" for sample in excludeSamples]
        excludeSamples = excludeSamples + ["fail","skip","unclassified","Undetermined","~$","pod5"]
        excludeSamples = "|".join(excludeSamples)

        # Move files
        fileCount = 0    

        for f in pathfilter:
            for p in glob.glob(f, recursive=True, root_dir=runDir):
                if os.path.isfile(os.path.join(runDir, p)) and not re.search(excludeSamples, p):   # Check if exists and isn't in the sample filter
                    p_dest = p if (group != "PulseNet") else os.path.basename(p) # Put in base of target directory if from Pulsenet. TODO: Do this in the pipeline script
                    os.makedirs(os.path.join(outDir, os.path.dirname(p_dest)), exist_ok=True)
                    src = os.path.join(runDir, p)
                    dst = os.path.join(outDir, p_dest)
                    if symlinkFQ and pathlib.Path(p).suffix.lower() in [".fastq", ".fq"]: # Symlink or not
                        os.symlink(src, dst)
                    else:
                        shutil.copy(src, dst)
                    fileCount = fileCount + 1
        printLog(f"      Copied files: {fileCount}")
        
        #subsetWorksheet(file, group, os.path.join(outDir,os.path.basename(file)))
        
    # Setup pipeline
    printLog(f"Configuring pipelines...")
    for group in groups:
        if group.lower() == "ignore": continue

        # Check for pipeline
        ignore = False
        try: ignore = pipelines[group].lower() == "ignore"
        except KeyError: ignore = True
        if (ignore): printLog(f"   Ignoring pipeline for {group}"); continue

        # Check output dir exists
        if not os.path.exists(directories[group]):
            printLog(f"   Output directory does not exist for {group}. Skipping.")
            continue

        printLog(f"   Generating SLURM for {group}...")

        # Parse controls
        samples = allSamples.loc[allSamples['Sample_Group'] == group]
        ctrls = samples.dropna(subset=['Control'])
        negCtrls = posCtrls = pd.DataFrame()
        samples = samples[samples['Control'].isna()]

        if(len(ctrls.index)):
            negCtrls = ctrls['Control'].str.lower() == "negative"
            negCtrls = ctrls.loc[negCtrls==True]
            negCtrls = ",".join(map(str,negCtrls[samplePosCol].values.tolist()))

            posCtrls = ctrls['Control'].str.lower() != "negative"
            posCtrls = ctrls.loc[posCtrls==True]
            posCtrls["Control"] = posCtrls[samplePosCol].astype(str) +","+ posCtrls["Control"].astype(str)
            posCtrls = " ".join(posCtrls["Control"].values.tolist())

        # Parse extra data
        accessions = allSamples.loc[allSamples['Sample_Group'] == group]

        # Create the SLURM command
        symlink = lambda dir,link: f"ln -s {st.findFiles2(os.path.join('./**',dir))[0]} {os.path.join(directories[group],link)}"

        def symlink(dir,link): # Creates symlink command
            path = os.path.join(directories[group],"**",dir)
            file = st.findFiles2(path)
            if (len(file) < 1): 
                raise Exception(f"Error: No directory found for '{path}'")
            elif(len(file) > 1):
                raise Exception(f"Error: More than 1 directory found for '{path}'")
            file = os.path.relpath(file[0],directories[group])
            link = os.path.join(directories[group],link)
            link = os.path.relpath(link,directories[group])
            return f"ln -s {file} {link}"

        commands = []
        if (group == "ncov" or group == "ncov-ww"): #TODO: Put this in pipeline script
            # Do symlinks
            if ("_I_" in runName):
                commands.append(symlink("Fastq","fastq")) 
            else:
                commands.append(symlink("fast5_pass","fast5"))   
                commands.append(symlink("fastq_pass","gup_out"))   

            # Go to parent dir
            parentDir = os.path.dirname(directories[group].rstrip("/")) + "/"
            commands.append("\ncd {}\n".format(parentDir))

            baseDir = os.path.basename(directories[group].strip("/"))

            # Run pipeline
            command = f"python {pipelines[group]} -d {parentDir} -r {baseDir} -b 2"
            if (group == "ncov-ww"): command = command + " -f"

            if len(posCtrls): command = f"{command} -p {posCtrls}"
            if len(negCtrls): command = f"{command} -c {negCtrls}"
            commands.append(command)

        elif(group == "ncov-R10"):
            command = f"bash {pipelines[group]} -r {directories[group]}"
            if len(posCtrls): command = f"{command} -p '{posCtrls}'"
            if len(negCtrls): command = f"{command} -c {negCtrls}"
            commands.append(command)

        elif (group == "PulseNet"): # TODO: Convert to own pipeline script
            parentDir = os.path.dirname(directories[group].rstrip("/")) + "/"
            baseDir = os.path.basename(directories[group].strip("/"))
            commands.append("conda activate pulsenet_analysis_pipeline")               
            commands.append(f"python {pipelines[group]} -d {parentDir} -r {baseDir}")

        elif (group != ""): # TODO: Pass the sample sheet to the pipelines instead of above code
            commands.append(f"bash {pipelines[group]} {directories[group]}")

        else:
            printLog(f"   No pipeline found for {group}.")
            continue

        # Generate the SLURM file
        SLURMfile = generateSLURM(SLURM = SLURM, 
                                jobName = group+"_"+runName, 
                                runName = runName, 
                                outputDir = directories[group], 
                                command = "\n".join(commands), 
                                email = email)
        out = subprocess.run(["sbatch",SLURMfile,"-v"], capture_output = True, text = True) # Launch the SLURM file
        printLog(f"      {out.stdout}")

    printLog(f"All files transferred and pipeline initialized\n")

# Import the arguments
parser = argparse.ArgumentParser(description='APL NGS Pipeline Launcher')
parser.add_argument("-r", "--run", help="Path to the run directory. Must contain the PipelineWorksheet.xlsx.", default = defaultSampleSheet)
parser.add_argument("-e", "--email", help="Notify status alerts by e-mail.", default = None)
parser.add_argument("-f", "--force", help="Will delete the target directories without notification.", action='store_true')
args = parser.parse_args()

runLauncher(args.run, None if args.email == "None" else args.email, args.force)